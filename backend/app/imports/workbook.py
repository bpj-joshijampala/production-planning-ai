from dataclasses import dataclass
from datetime import date, datetime
from hashlib import sha256
from io import BytesIO
import json
import re
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

REQUIRED_SHEETS = ("Valve_Plan", "Component_Status", "Routing_Master", "Machine_Master", "Vendor_Master")

HEADER_ALIASES = {
    "alt_machine": "alt_machine",
    "assembly_date": "assembly_date",
    "buffer_days": "buffer_days",
    "capability_notes": "capability_notes",
    "capacity_rating": "capacity_rating",
    "comments": "comments",
    "component": "component",
    "component_line_no": "component_line_no",
    "critical": "critical",
    "current_location": "current_location",
    "customer": "customer",
    "description": "description",
    "dispatch_date": "dispatch_date",
    "effective_hours_day": "effective_hours_day",
    "effective_lead_days": "effective_lead_days",
    "efficiency_percent": "efficiency_percent",
    "efficiency": "efficiency_percent",
    "expected_from_fabrication": "expected_from_fabrication",
    "expected_ready_date": "expected_ready_date",
    "fabrication_complete": "fabrication_complete",
    "fabrication_required": "fabrication_required",
    "hours_per_day": "hours_per_day",
    "machine_id": "machine_id",
    "machine_type": "machine_type",
    "notes": "notes",
    "operation_name": "operation_name",
    "operation_no": "operation_no",
    "order_id": "order_id",
    "primary_process": "primary_process",
    "priority": "priority",
    "priority_eligible": "priority_eligible",
    "qty": "qty",
    "ready_date_type": "ready_date_type",
    "reliability": "reliability",
    "remarks": "remarks",
    "shift_pattern": "shift_pattern",
    "status": "status",
    "std_run_hrs": "std_run_hrs",
    "std_setup_hrs": "std_setup_hrs",
    "std_total_hrs": "std_total_hrs",
    "subcontract_allowed": "subcontract_allowed",
    "transport_days_total": "transport_days_total",
    "turnaround_days": "turnaround_days",
    "value_cr": "value_cr",
    "value": "value_cr",
    "valve_id": "valve_id",
    "valve_type": "valve_type",
    "vendor_id": "vendor_id",
    "vendor_name": "vendor_name",
    "vendor_process": "vendor_process",
}


class WorkbookParseError(ValueError):
    pass


@dataclass(frozen=True)
class ParsedWorkbookRow:
    sheet_name: str
    row_number: int
    payload: dict[str, Any]
    row_hash: str


@dataclass(frozen=True)
class ParsedWorkbook:
    workbook_sheet_names: tuple[str, ...]
    headers_by_sheet: dict[str, tuple[str, ...]]
    rows: list[ParsedWorkbookRow]


def parse_workbook(content: bytes) -> list[ParsedWorkbookRow]:
    return parse_workbook_with_metadata(content).rows


def parse_workbook_with_metadata(content: bytes) -> ParsedWorkbook:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
        raise WorkbookParseError("Uploaded file is not a readable .xlsx workbook.") from exc

    parsed_rows: list[ParsedWorkbookRow] = []
    headers_by_sheet: dict[str, tuple[str, ...]] = {}
    for sheet_name in REQUIRED_SHEETS:
        if sheet_name not in workbook.sheetnames:
            continue

        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        raw_headers = next(rows, None)
        if raw_headers is None:
            headers_by_sheet[sheet_name] = ()
            continue

        headers = [normalize_header(header) if header is not None else "" for header in raw_headers]
        headers_by_sheet[sheet_name] = tuple(header for header in headers if header)
        for row_number, values in enumerate(rows, start=2):
            if _row_is_blank(values):
                continue

            payload = _payload_from_row(headers, values)
            parsed_rows.append(
                ParsedWorkbookRow(
                    sheet_name=sheet_name,
                    row_number=row_number,
                    payload=payload,
                    row_hash=hash_payload(payload),
                )
            )

    workbook_sheet_names = tuple(workbook.sheetnames)
    workbook.close()
    return ParsedWorkbook(workbook_sheet_names=workbook_sheet_names, headers_by_sheet=headers_by_sheet, rows=parsed_rows)


def hash_payload(payload: dict[str, Any]) -> str:
    payload_json = json.dumps(payload, sort_keys=True, separators=(",", ":"), ensure_ascii=True)
    return sha256(payload_json.encode("utf-8")).hexdigest()


def normalize_header(header: Any) -> str:
    cleaned = str(header).strip().lower()
    cleaned = cleaned.replace("%", " percent ")
    cleaned = re.sub(r"[^a-z0-9]+", " ", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    key = cleaned.replace(" ", "_")
    return HEADER_ALIASES.get(key, key)


def _payload_from_row(headers: list[str], values: tuple[Any, ...]) -> dict[str, Any]:
    payload: dict[str, Any] = {}
    for header, value in zip(headers, values, strict=False):
        if not header:
            continue
        payload[header] = _json_safe_value(value)
    return payload


def _json_safe_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def _row_is_blank(values: tuple[Any, ...]) -> bool:
    return all(value is None or str(value).strip() == "" for value in values)
