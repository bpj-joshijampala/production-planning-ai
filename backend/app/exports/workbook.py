from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from datetime import date, datetime
import json

from openpyxl import Workbook


@dataclass(frozen=True)
class ExportSheet:
    name: str
    columns: tuple[str, ...]
    rows: tuple[Mapping[str, object | None], ...]


def build_export_workbook(
    *,
    export_info_rows: Sequence[tuple[str, object | None]],
    sheets: Sequence[ExportSheet],
) -> Workbook:
    if not sheets:
        raise ValueError("At least one export sheet is required.")

    workbook = Workbook()
    workbook.remove(workbook.active)

    info_sheet = workbook.create_sheet(title="Export_Info")
    info_sheet.append(("Field", "Value"))
    for field, value in export_info_rows:
        info_sheet.append((field, _cell_value(value)))

    used_sheet_names = {"Export_Info"}
    for sheet in sheets:
        title = sheet.name.strip()
        if not title:
            raise ValueError("Export sheet name must not be blank.")
        if len(title) > 31:
            raise ValueError(f"Export sheet name {title!r} exceeds Excel's 31 character limit.")
        if title in used_sheet_names:
            raise ValueError(f"Duplicate export sheet name {title!r}.")
        if not sheet.columns:
            raise ValueError(f"Export sheet {title!r} must define at least one column.")

        used_sheet_names.add(title)
        worksheet = workbook.create_sheet(title=title)
        worksheet.append(list(sheet.columns))
        for row_index, row in enumerate(sheet.rows, start=1):
            missing_columns = [column for column in sheet.columns if column not in row]
            if missing_columns:
                raise ValueError(
                    f"Export sheet {title!r} row {row_index} is missing required columns: {', '.join(missing_columns)}."
                )
            worksheet.append([_cell_value(row[column]) for column in sheet.columns])

    return workbook


def _cell_value(value: object | None) -> object | None:
    if value is None or isinstance(value, str | int | float | bool):
        return value
    if isinstance(value, date | datetime):
        return value.isoformat()
    if isinstance(value, Mapping):
        return json.dumps(dict(value), sort_keys=True, separators=(",", ":"), ensure_ascii=True)
    if isinstance(value, Sequence) and not isinstance(value, str | bytes | bytearray):
        return json.dumps(list(value), separators=(",", ":"), ensure_ascii=True)
    return str(value)
