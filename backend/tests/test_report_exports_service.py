import json
from pathlib import Path
from copy import deepcopy

from alembic import command
from alembic.config import Config
from fastapi import HTTPException
from fastapi.testclient import TestClient
from openpyxl import load_workbook
import pytest
from sqlalchemy import func, select

from app.core.config import get_settings
from app.db.session import create_session_factory
from app.exports.workbook import ExportSheet
from app.main import create_app
from app.models.output import (
    FlowBlocker,
    MachineLoadSummary,
    PlannedOperation,
    Recommendation,
    ReportExport,
    ValveReadinessSummary,
)
from app.services.report_exports import generate_first_build_report_export, generate_xlsx_report_export
from tests.workbook_fixtures import minimal_workbook_rows, workbook_bytes


DEV_USER_ID = "00000000-0000-0000-0000-000000000001"


@pytest.fixture(name="client")
def fixture_client(tmp_path, monkeypatch):  # type: ignore[no-untyped-def]
    database_path = tmp_path / "report_exports_service.sqlite3"
    upload_dir = tmp_path / "uploads"
    export_dir = tmp_path / "exports"

    monkeypatch.setenv("DATABASE_URL", f"sqlite:///{database_path.as_posix()}")
    monkeypatch.setenv("UPLOAD_DIR", upload_dir.as_posix())
    monkeypatch.setenv("EXPORT_DIR", export_dir.as_posix())
    get_settings.cache_clear()

    command.upgrade(Config("alembic.ini"), "head")

    with TestClient(create_app()) as test_client:
        yield test_client

    get_settings.cache_clear()


def test_generate_xlsx_report_export_creates_workbook_and_audit_record(client: TestClient) -> None:
    planning_run_id = _create_calculated_planning_run(client)

    session_factory = create_session_factory()
    with session_factory() as session:
        machine_rows = [
            {
                "Machine_Type": row.machine_type,
                "Total_Operation_Hours": row.total_operation_hours,
                "Status": row.status,
            }
            for row in session.scalars(
                select(MachineLoadSummary)
                .where(MachineLoadSummary.planning_run_id == planning_run_id)
                .order_by(MachineLoadSummary.machine_type.asc())
            )
        ]
        report_export = generate_xlsx_report_export(
            planning_run_id=planning_run_id,
            report_type="MACHINE_LOAD",
            generated_by_user_id=DEV_USER_ID,
            sheets=(
                ExportSheet(
                    name="Machine_Load",
                    columns=("Machine_Type", "Total_Operation_Hours", "Status"),
                    rows=tuple(machine_rows),
                ),
            ),
            db=session,
        )

    export_path = Path(report_export.file_path)
    assert export_path.exists()
    assert report_export.file_format == "XLSX"
    assert report_export.generated_by_user_id == DEV_USER_ID

    metadata = json.loads(report_export.metadata_json or "{}")
    assert metadata == {
        "sheet_names": ["Machine_Load"],
        "sheet_row_counts": {"Machine_Load": 1},
    }

    workbook = load_workbook(export_path, data_only=True)
    assert workbook.sheetnames == ["Export_Info", "Machine_Load"]

    export_info_rows = {
        str(field): value
        for field, value in workbook["Export_Info"].iter_rows(min_row=2, values_only=True)
    }
    assert export_info_rows["Report_Type"] == "MACHINE_LOAD"
    assert export_info_rows["PlanningRun_ID"] == planning_run_id
    assert export_info_rows["Upload_File"] == "plan.xlsx"
    assert export_info_rows["Planning_Start_Date"] == "2026-04-21"
    assert export_info_rows["Planning_Horizon_Days"] == 7
    assert export_info_rows["Generated_At"] == report_export.generated_at
    assert export_info_rows["Generated_By"] == "Development Planner"

    machine_sheet = workbook["Machine_Load"]
    header = [cell.value for cell in machine_sheet[1]]
    first_row = [cell.value for cell in machine_sheet[2]]
    assert header == ["Machine_Type", "Total_Operation_Hours", "Status"]
    assert first_row == [
        machine_rows[0]["Machine_Type"],
        machine_rows[0]["Total_Operation_Hours"],
        machine_rows[0]["Status"],
    ]

    with session_factory() as session:
        persisted = session.get(ReportExport, report_export.id)

    assert persisted is not None
    assert persisted.report_type == "MACHINE_LOAD"
    assert persisted.file_path == str(export_path)


def test_generate_xlsx_report_export_removes_file_if_audit_write_fails(
    client: TestClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    planning_run_id = _create_calculated_planning_run(client)

    session_factory = create_session_factory()
    with session_factory() as session:
        machine_rows = [
            {
                "Machine_Type": row.machine_type,
                "Total_Operation_Hours": row.total_operation_hours,
                "Status": row.status,
            }
            for row in session.scalars(
                select(MachineLoadSummary)
                .where(MachineLoadSummary.planning_run_id == planning_run_id)
                .order_by(MachineLoadSummary.machine_type.asc())
            )
        ]

        def fail_commit() -> None:
            raise RuntimeError("audit commit failed")

        monkeypatch.setattr(session, "commit", fail_commit)

        export_dir = get_settings().export_dir / planning_run_id
        with pytest.raises(RuntimeError, match="audit commit failed"):
            generate_xlsx_report_export(
                planning_run_id=planning_run_id,
                report_type="MACHINE_LOAD",
                generated_by_user_id=DEV_USER_ID,
                sheets=(
                    ExportSheet(
                        name="Machine_Load",
                        columns=("Machine_Type", "Total_Operation_Hours", "Status"),
                        rows=tuple(machine_rows),
                    ),
                ),
                db=session,
            )

    assert not export_dir.exists() or not list(export_dir.glob("*.xlsx"))

    with session_factory() as session:
        export_count = session.scalar(
            select(func.count()).select_from(ReportExport).where(ReportExport.planning_run_id == planning_run_id)
        )

    assert export_count == 0


def test_generate_xlsx_report_export_keeps_file_if_refresh_fails_after_audit_commit(
    client: TestClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    planning_run_id = _create_calculated_planning_run(client)
    export_dir = get_settings().export_dir / planning_run_id

    session_factory = create_session_factory()
    with session_factory() as session:
        machine_rows = [
            {
                "Machine_Type": row.machine_type,
                "Total_Operation_Hours": row.total_operation_hours,
                "Status": row.status,
            }
            for row in session.scalars(
                select(MachineLoadSummary)
                .where(MachineLoadSummary.planning_run_id == planning_run_id)
                .order_by(MachineLoadSummary.machine_type.asc())
            )
        ]

        def fail_refresh(_instance: object) -> None:
            raise RuntimeError("post-commit refresh failed")

        monkeypatch.setattr(session, "refresh", fail_refresh)

        with pytest.raises(RuntimeError, match="post-commit refresh failed"):
            generate_xlsx_report_export(
                planning_run_id=planning_run_id,
                report_type="MACHINE_LOAD",
                generated_by_user_id=DEV_USER_ID,
                sheets=(
                    ExportSheet(
                        name="Machine_Load",
                        columns=("Machine_Type", "Total_Operation_Hours", "Status"),
                        rows=tuple(machine_rows),
                    ),
                ),
                db=session,
            )

    generated_files = list(export_dir.glob("*.xlsx"))
    assert len(generated_files) == 1

    with session_factory() as session:
        persisted_exports = list(
            session.scalars(select(ReportExport).where(ReportExport.planning_run_id == planning_run_id))
        )

    assert len(persisted_exports) == 1
    assert persisted_exports[0].file_path == str(generated_files[0])
    assert generated_files[0].exists()


def test_generate_xlsx_report_export_uses_unique_file_names_when_timestamp_matches(
    client: TestClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    planning_run_id = _create_calculated_planning_run(client)
    monkeypatch.setattr("app.services.report_exports.utc_now_iso", lambda: "2026-05-01T09:00:00.000000Z")

    session_factory = create_session_factory()
    with session_factory() as session:
        first_export = generate_first_build_report_export(
            planning_run_id=planning_run_id,
            report_type="MACHINE_LOAD",
            file_format="XLSX",
            db=session,
        )
        second_export = generate_first_build_report_export(
            planning_run_id=planning_run_id,
            report_type="MACHINE_LOAD",
            file_format="XLSX",
            db=session,
        )
        first_export_id = first_export.id
        second_export_id = second_export.id
        first_path = Path(first_export.file_path)
        second_path = Path(second_export.file_path)

    assert first_export_id != second_export_id
    assert first_path != second_path
    assert first_path.exists()
    assert second_path.exists()
    assert first_path.name.startswith("machine_load_20260501T090000000000Z_")
    assert second_path.name.startswith("machine_load_20260501T090000000000Z_")


def test_generate_xlsx_report_export_rejects_non_calculated_run(client: TestClient) -> None:
    upload_response = client.post(
        "/api/v1/uploads",
        files={
            "file": (
                "plan.xlsx",
                workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert upload_response.status_code == 201

    planning_run_response = client.post(
        "/api/v1/planning-runs",
        json={
            "upload_batch_id": upload_response.json()["id"],
            "planning_start_date": "2026-04-21",
            "planning_horizon_days": 7,
        },
    )
    assert planning_run_response.status_code == 201
    planning_run_id = str(planning_run_response.json()["id"])

    session_factory = create_session_factory()
    with session_factory() as session:
        with pytest.raises(HTTPException) as exc_info:
            generate_xlsx_report_export(
                planning_run_id=planning_run_id,
                report_type="MACHINE_LOAD",
                generated_by_user_id=DEV_USER_ID,
                sheets=(
                    ExportSheet(
                        name="Machine_Load",
                        columns=("Machine_Type", "Total_Operation_Hours", "Status"),
                        rows=(),
                    ),
                ),
                db=session,
            )

    response = exc_info.value
    assert response.status_code == 409
    assert response.detail["code"] == "PLANNING_RUN_NOT_CALCULATED"


def test_generate_xlsx_report_export_removes_file_if_workbook_save_fails(
    client: TestClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    planning_run_id = _create_calculated_planning_run(client)

    session_factory = create_session_factory()
    with session_factory() as session:
        machine_rows = [
            {
                "Machine_Type": row.machine_type,
                "Total_Operation_Hours": row.total_operation_hours,
                "Status": row.status,
            }
            for row in session.scalars(
                select(MachineLoadSummary)
                .where(MachineLoadSummary.planning_run_id == planning_run_id)
                .order_by(MachineLoadSummary.machine_type.asc())
            )
        ]

        def fail_save(*_args: object, **_kwargs: object) -> None:
            raise RuntimeError("workbook save failed")

        monkeypatch.setattr("app.exports.workbook.Workbook.save", fail_save)

        export_dir = get_settings().export_dir / planning_run_id
        with pytest.raises(RuntimeError, match="workbook save failed"):
            generate_xlsx_report_export(
                planning_run_id=planning_run_id,
                report_type="MACHINE_LOAD",
                generated_by_user_id=DEV_USER_ID,
                sheets=(
                    ExportSheet(
                        name="Machine_Load",
                        columns=("Machine_Type", "Total_Operation_Hours", "Status"),
                        rows=tuple(machine_rows),
                    ),
                ),
                db=session,
            )

    assert not export_dir.exists() or not list(export_dir.glob("*.xlsx"))

    with session_factory() as session:
        export_count = session.scalar(
            select(func.count()).select_from(ReportExport).where(ReportExport.planning_run_id == planning_run_id)
        )

    assert export_count == 0


def test_build_export_workbook_fails_on_missing_required_column() -> None:
    from app.exports.workbook import build_export_workbook

    with pytest.raises(ValueError, match="missing required columns: Status"):
        build_export_workbook(
            export_info_rows=(("Report_Type", "MACHINE_LOAD"),),
            sheets=(
                ExportSheet(
                    name="Machine_Load",
                    columns=("Machine_Type", "Status"),
                    rows=({"Machine_Type": "HBM"},),
                ),
            ),
        )


@pytest.mark.parametrize(
    ("report_type", "sheet_name", "expected_headers"),
    [
        (
            "MACHINE_LOAD",
            "Machine_Load",
            [
                "Machine_Type",
                "Total_Operation_Hours",
                "Capacity_Hours_Per_Day",
                "Load_Days",
                "Buffer_Days",
                "Overload_Flag",
                "Overload_Days",
                "Spare_Capacity_Days",
                "Underutilized_Flag",
                "Batch_Risk_Flag",
                "Status",
            ],
        ),
        (
            "SUBCONTRACT_PLAN",
            "Subcontract_Plan",
            [
                "Recommendation_Type",
                "Valve_ID",
                "Component_Line_No",
                "Component",
                "Operation_Name",
                "Machine_Type",
                "Suggested_Vendor_ID",
                "Suggested_Vendor_Name",
                "Internal_Wait_Days",
                "Internal_Completion_Days",
                "Vendor_Total_Days",
                "Vendor_Gain_Days",
                "Batch_Candidate_Count",
                "Batch_Opportunity",
                "Status",
                "Explanation",
            ],
        ),
        (
            "VALVE_READINESS",
            "Valve_Readiness",
            [
                "Valve_ID",
                "Customer",
                "Assembly_Date",
                "Dispatch_Date",
                "Value_Cr",
                "Total_Components",
                "Ready_Components",
                "Required_Components",
                "Ready_Required_Count",
                "Pending_Required_Count",
                "Full_Kit",
                "Near_Ready",
                "Expected_Completion_Date",
                "Assembly_Delay_Days",
                "Status",
                "Risk_Reason",
                "Valve_Flow_Gap_Days",
                "Valve_Flow_Imbalance",
            ],
        ),
        (
            "FLOW_BLOCKER",
            "Flow_Blockers",
            [
                "Severity",
                "Blocker_Type",
                "Valve_ID",
                "Component_Line_No",
                "Component",
                "Operation_Name",
                "Cause",
                "Recommended_Action",
            ],
        ),
        (
            "DAILY_EXECUTION",
            "Daily_Execution",
            [
                "Date",
                "Machine_Type",
                "Queue_Sequence",
                "Valve_ID",
                "Component_Line_No",
                "Component",
                "Operation_Name",
                "Planned_Action",
                "Internal_Wait_Days",
                "Internal_Completion_Date",
                "Extreme_Delay_Flag",
            ],
        ),
    ],
)
def test_generate_first_build_report_export_creates_expected_report_sheet(
    client: TestClient,
    report_type: str,
    sheet_name: str,
    expected_headers: list[str],
) -> None:
    planning_run_id = _create_calculated_planning_run(
        client,
        workbook_content=workbook_bytes(sheets=_first_build_export_workbook_rows()),
    )

    session_factory = create_session_factory()
    with session_factory() as session:
        report_export = generate_first_build_report_export(
            planning_run_id=planning_run_id,
            report_type=report_type,
            file_format="XLSX",
            db=session,
        )

    workbook = load_workbook(report_export.file_path, data_only=True)
    assert workbook.sheetnames == ["Export_Info", sheet_name]
    worksheet = workbook[sheet_name]
    assert [cell.value for cell in worksheet[1]] == expected_headers
    actual_rows = _worksheet_records(worksheet)
    assert actual_rows == _expected_first_build_report_rows(
        planning_run_id=planning_run_id,
        report_type=report_type,
    )

    if report_type == "SUBCONTRACT_PLAN":
        assert {row["Recommendation_Type"] for row in actual_rows} <= {
            "SUBCONTRACT",
            "BATCH_SUBCONTRACT_OPPORTUNITY",
        }
        assert any(row["Batch_Opportunity"] == 1 for row in actual_rows)


def test_first_build_report_exports_reflect_current_database_output_values(client: TestClient) -> None:
    planning_run_id = _create_calculated_planning_run(
        client,
        workbook_content=workbook_bytes(sheets=_first_build_export_workbook_rows()),
    )

    session_factory = create_session_factory()
    with session_factory() as session:
        machine_load = session.scalar(
            select(MachineLoadSummary)
            .where(MachineLoadSummary.planning_run_id == planning_run_id)
            .where(MachineLoadSummary.machine_type == "HBM")
        )
        assert machine_load is not None
        machine_load.total_operation_hours = 123.45
        machine_load.capacity_hours_per_day = 6.5
        machine_load.load_days = 18.99
        machine_load.buffer_days = 4.5
        machine_load.overload_flag = 1
        machine_load.overload_days = 14.49
        machine_load.spare_capacity_days = 0.0
        machine_load.underutilized_flag = 0
        machine_load.batch_risk_flag = 1
        machine_load.status = "OVERLOADED"

        readiness = session.scalar(
            select(ValveReadinessSummary)
            .where(ValveReadinessSummary.planning_run_id == planning_run_id)
            .where(ValveReadinessSummary.valve_id == "V-100")
        )
        assert readiness is not None
        readiness.otd_delay_days = 7.25
        readiness.readiness_status = "AT_RISK"
        readiness.risk_reason = "Sentinel assembly risk"
        readiness.valve_flow_gap_days = 3.5
        readiness.valve_flow_imbalance_flag = 1

        operation = session.scalar(
            select(PlannedOperation)
            .where(PlannedOperation.planning_run_id == planning_run_id)
            .where(PlannedOperation.valve_id == "V-100")
            .where(PlannedOperation.component_line_no == 1)
            .order_by(PlannedOperation.operation_no.asc())
        )
        assert operation is not None
        operation.operation_name = "Sentinel daily op"
        operation.operation_arrival_date = "2026-04-23"
        operation.internal_wait_days = 4.75
        operation.internal_completion_date = "2026-04-29"
        operation.extreme_delay_flag = 1

        session.add(
            Recommendation(
                id="sentinel-subcontract-recommendation",
                planning_run_id=planning_run_id,
                planned_operation_id=None,
                recommendation_type="SUBCONTRACT",
                valve_id="V-100",
                component_line_no=1,
                component="Body",
                operation_name="Sentinel subcontract op",
                machine_type="HBM",
                suggested_machine_type=None,
                suggested_vendor_id="VEN-SENTINEL",
                suggested_vendor_name="Sentinel Vendor",
                internal_wait_days=8.5,
                processing_time_days=None,
                internal_completion_days=12.5,
                vendor_total_days=2.0,
                vendor_gain_days=10.5,
                subcontract_batch_candidate_count=2,
                batch_subcontract_opportunity_flag=1,
                reason_codes_json='["SENTINEL_SUBCONTRACT"]',
                explanation="Sentinel subcontract recommendation",
                status="PENDING",
                created_at="2998-01-01T00:00:00Z",
            )
        )
        session.add(
            Recommendation(
                id="sentinel-daily-action",
                planning_run_id=planning_run_id,
                planned_operation_id=operation.id,
                recommendation_type="USE_ALTERNATE",
                valve_id=operation.valve_id,
                component_line_no=operation.component_line_no,
                component=operation.component,
                operation_name=operation.operation_name,
                machine_type=operation.machine_type,
                suggested_machine_type="VTL",
                suggested_vendor_id=None,
                suggested_vendor_name=None,
                internal_wait_days=operation.internal_wait_days,
                processing_time_days=operation.processing_time_days,
                internal_completion_days=operation.internal_completion_days,
                vendor_total_days=None,
                vendor_gain_days=None,
                subcontract_batch_candidate_count=None,
                batch_subcontract_opportunity_flag=0,
                reason_codes_json='["SENTINEL_ACTION"]',
                explanation="Sentinel daily execution action",
                status="PENDING",
                created_at="2999-01-01T00:00:00Z",
            )
        )
        session.add(
            FlowBlocker(
                id="sentinel-flow-blocker",
                planning_run_id=planning_run_id,
                planned_operation_id=operation.id,
                valve_id="V-100",
                component_line_no=1,
                component="Body",
                operation_name="Sentinel blocked op",
                blocker_type="MISSING_MACHINE",
                cause="Sentinel blocker cause",
                recommended_action="Sentinel blocker action",
                severity="CRITICAL",
                created_at="2999-01-01T00:00:00Z",
            )
        )
        session.commit()

    machine_rows = _generated_report_rows(planning_run_id=planning_run_id, report_type="MACHINE_LOAD")
    assert _row_by(machine_rows, "Machine_Type", "HBM") == {
        "Machine_Type": "HBM",
        "Total_Operation_Hours": 123.45,
        "Capacity_Hours_Per_Day": 6.5,
        "Load_Days": 18.99,
        "Buffer_Days": 4.5,
        "Overload_Flag": 1,
        "Overload_Days": 14.49,
        "Spare_Capacity_Days": 0.0,
        "Underutilized_Flag": 0,
        "Batch_Risk_Flag": 1,
        "Status": "OVERLOADED",
    }

    subcontract_rows = _generated_report_rows(planning_run_id=planning_run_id, report_type="SUBCONTRACT_PLAN")
    assert _row_by(subcontract_rows, "Explanation", "Sentinel subcontract recommendation") == {
        "Recommendation_Type": "SUBCONTRACT",
        "Valve_ID": "V-100",
        "Component_Line_No": 1,
        "Component": "Body",
        "Operation_Name": "Sentinel subcontract op",
        "Machine_Type": "HBM",
        "Suggested_Vendor_ID": "VEN-SENTINEL",
        "Suggested_Vendor_Name": "Sentinel Vendor",
        "Internal_Wait_Days": 8.5,
        "Internal_Completion_Days": 12.5,
        "Vendor_Total_Days": 2.0,
        "Vendor_Gain_Days": 10.5,
        "Batch_Candidate_Count": 2,
        "Batch_Opportunity": 1,
        "Status": "PENDING",
        "Explanation": "Sentinel subcontract recommendation",
    }

    readiness_rows = _generated_report_rows(planning_run_id=planning_run_id, report_type="VALVE_READINESS")
    readiness_row = _row_by(readiness_rows, "Valve_ID", "V-100")
    assert readiness_row["Assembly_Delay_Days"] == 7.25
    assert readiness_row["Status"] == "AT_RISK"
    assert readiness_row["Risk_Reason"] == "Sentinel assembly risk"
    assert readiness_row["Valve_Flow_Gap_Days"] == 3.5
    assert readiness_row["Valve_Flow_Imbalance"] == 1

    blocker_rows = _generated_report_rows(planning_run_id=planning_run_id, report_type="FLOW_BLOCKER")
    assert _row_by(blocker_rows, "Cause", "Sentinel blocker cause") == {
        "Severity": "CRITICAL",
        "Blocker_Type": "MISSING_MACHINE",
        "Valve_ID": "V-100",
        "Component_Line_No": 1,
        "Component": "Body",
        "Operation_Name": "Sentinel blocked op",
        "Cause": "Sentinel blocker cause",
        "Recommended_Action": "Sentinel blocker action",
    }

    execution_rows = _generated_report_rows(planning_run_id=planning_run_id, report_type="DAILY_EXECUTION")
    execution_row = _row_by(execution_rows, "Operation_Name", "Sentinel daily op")
    assert execution_row["Date"] == "2026-04-23"
    assert execution_row["Planned_Action"] == "USE_ALTERNATE"
    assert execution_row["Internal_Wait_Days"] == 4.75
    assert execution_row["Internal_Completion_Date"] == "2026-04-29"
    assert execution_row["Extreme_Delay_Flag"] == 1


def _worksheet_records(worksheet) -> list[dict[str, object | None]]:  # type: ignore[no-untyped-def]
    headers = [cell.value for cell in worksheet[1]]
    return [
        dict(zip(headers, row_values, strict=True))
        for row_values in worksheet.iter_rows(min_row=2, values_only=True)
    ]


def _generated_report_rows(*, planning_run_id: str, report_type: str) -> list[dict[str, object | None]]:
    session_factory = create_session_factory()
    with session_factory() as session:
        report_export = generate_first_build_report_export(
            planning_run_id=planning_run_id,
            report_type=report_type,
            file_format="XLSX",
            db=session,
        )

    workbook = load_workbook(report_export.file_path, data_only=True)
    return _worksheet_records(workbook[workbook.sheetnames[1]])


def _row_by(
    rows: list[dict[str, object | None]],
    column: str,
    value: object,
) -> dict[str, object | None]:
    for row in rows:
        if row[column] == value:
            return row
    raise AssertionError(f"Could not find row where {column} == {value!r}. Rows: {rows!r}")


def _expected_first_build_report_rows(
    *,
    planning_run_id: str,
    report_type: str,
) -> list[dict[str, object | None]]:
    session_factory = create_session_factory()
    with session_factory() as session:
        if report_type == "MACHINE_LOAD":
            return [
                {
                    "Machine_Type": row.machine_type,
                    "Total_Operation_Hours": row.total_operation_hours,
                    "Capacity_Hours_Per_Day": row.capacity_hours_per_day,
                    "Load_Days": row.load_days,
                    "Buffer_Days": row.buffer_days,
                    "Overload_Flag": row.overload_flag,
                    "Overload_Days": row.overload_days,
                    "Spare_Capacity_Days": row.spare_capacity_days,
                    "Underutilized_Flag": row.underutilized_flag,
                    "Batch_Risk_Flag": row.batch_risk_flag,
                    "Status": row.status,
                }
                for row in session.scalars(
                    select(MachineLoadSummary)
                    .where(MachineLoadSummary.planning_run_id == planning_run_id)
                    .order_by(MachineLoadSummary.machine_type.asc(), MachineLoadSummary.id.asc())
                )
            ]

        if report_type == "SUBCONTRACT_PLAN":
            return [
                {
                    "Recommendation_Type": row.recommendation_type,
                    "Valve_ID": row.valve_id,
                    "Component_Line_No": row.component_line_no,
                    "Component": row.component,
                    "Operation_Name": row.operation_name,
                    "Machine_Type": row.machine_type,
                    "Suggested_Vendor_ID": row.suggested_vendor_id,
                    "Suggested_Vendor_Name": row.suggested_vendor_name,
                    "Internal_Wait_Days": row.internal_wait_days,
                    "Internal_Completion_Days": row.internal_completion_days,
                    "Vendor_Total_Days": row.vendor_total_days,
                    "Vendor_Gain_Days": row.vendor_gain_days,
                    "Batch_Candidate_Count": row.subcontract_batch_candidate_count,
                    "Batch_Opportunity": row.batch_subcontract_opportunity_flag,
                    "Status": row.status,
                    "Explanation": row.explanation,
                }
                for row in session.scalars(
                    select(Recommendation)
                    .where(Recommendation.planning_run_id == planning_run_id)
                    .where(
                        Recommendation.recommendation_type.in_(
                            ("SUBCONTRACT", "BATCH_SUBCONTRACT_OPPORTUNITY")
                        )
                    )
                    .order_by(
                        Recommendation.suggested_vendor_id.asc(),
                        Recommendation.valve_id.asc(),
                        Recommendation.component_line_no.asc(),
                        Recommendation.operation_name.asc(),
                        Recommendation.id.asc(),
                    )
                )
            ]

        if report_type == "VALVE_READINESS":
            return [
                {
                    "Valve_ID": row.valve_id,
                    "Customer": row.customer,
                    "Assembly_Date": row.assembly_date,
                    "Dispatch_Date": row.dispatch_date,
                    "Value_Cr": row.value_cr,
                    "Total_Components": row.total_components,
                    "Ready_Components": row.ready_components,
                    "Required_Components": row.required_components,
                    "Ready_Required_Count": row.ready_required_count,
                    "Pending_Required_Count": row.pending_required_count,
                    "Full_Kit": row.full_kit_flag,
                    "Near_Ready": row.near_ready_flag,
                    "Expected_Completion_Date": row.valve_expected_completion_date,
                    "Assembly_Delay_Days": row.otd_delay_days,
                    "Status": row.readiness_status,
                    "Risk_Reason": row.risk_reason,
                    "Valve_Flow_Gap_Days": row.valve_flow_gap_days,
                    "Valve_Flow_Imbalance": row.valve_flow_imbalance_flag,
                }
                for row in session.scalars(
                    select(ValveReadinessSummary)
                    .where(ValveReadinessSummary.planning_run_id == planning_run_id)
                    .order_by(ValveReadinessSummary.assembly_date.asc(), ValveReadinessSummary.valve_id.asc())
                )
            ]

        if report_type == "FLOW_BLOCKER":
            severity_rank = {"CRITICAL": 0, "WARNING": 1, "INFO": 2}
            rows = list(
                session.scalars(select(FlowBlocker).where(FlowBlocker.planning_run_id == planning_run_id))
            )
            rows.sort(
                key=lambda row: (
                    severity_rank.get(row.severity, 99),
                    row.blocker_type or "",
                    row.valve_id or "",
                    row.component_line_no or 0,
                    row.operation_name or "",
                    row.id,
                )
            )
            return [
                {
                    "Severity": row.severity,
                    "Blocker_Type": row.blocker_type,
                    "Valve_ID": row.valve_id,
                    "Component_Line_No": row.component_line_no,
                    "Component": row.component,
                    "Operation_Name": row.operation_name,
                    "Cause": row.cause,
                    "Recommended_Action": row.recommended_action,
                }
                for row in rows
            ]

        if report_type == "DAILY_EXECUTION":
            operations = list(
                session.scalars(select(PlannedOperation).where(PlannedOperation.planning_run_id == planning_run_id))
            )
            operations.sort(
                key=lambda row: (
                    row.internal_completion_date is None,
                    row.internal_completion_date or "",
                    row.sort_sequence,
                    row.id,
                )
            )
            action_by_operation_id: dict[str, str] = {}
            for row in session.scalars(
                select(Recommendation)
                .where(Recommendation.planning_run_id == planning_run_id)
                .where(Recommendation.planned_operation_id.is_not(None))
                .order_by(Recommendation.created_at.desc(), Recommendation.id.desc())
            ):
                if row.planned_operation_id is not None and row.planned_operation_id not in action_by_operation_id:
                    action_by_operation_id[row.planned_operation_id] = row.recommendation_type

            return [
                {
                    "Date": row.operation_arrival_date,
                    "Machine_Type": row.machine_type,
                    "Queue_Sequence": row.sort_sequence,
                    "Valve_ID": row.valve_id,
                    "Component_Line_No": row.component_line_no,
                    "Component": row.component,
                    "Operation_Name": row.operation_name,
                    "Planned_Action": action_by_operation_id.get(row.id, "OK_INTERNAL"),
                    "Internal_Wait_Days": row.internal_wait_days,
                    "Internal_Completion_Date": row.internal_completion_date,
                    "Extreme_Delay_Flag": row.extreme_delay_flag,
                }
                for row in operations
            ]

    raise AssertionError(f"Unexpected report type {report_type}.")


def _create_calculated_planning_run(client: TestClient, workbook_content: bytes | None = None) -> str:
    upload_response = client.post(
        "/api/v1/uploads",
        files={
            "file": (
                "plan.xlsx",
                workbook_content if workbook_content is not None else workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert upload_response.status_code == 201

    planning_run_response = client.post(
        "/api/v1/planning-runs",
        json={
            "upload_batch_id": upload_response.json()["id"],
            "planning_start_date": "2026-04-21",
            "planning_horizon_days": 7,
        },
    )
    assert planning_run_response.status_code == 201

    planning_run_id = str(planning_run_response.json()["id"])
    calculate_response = client.post(f"/api/v1/planning-runs/{planning_run_id}/calculate")
    assert calculate_response.status_code == 200
    return planning_run_id


def _first_build_export_workbook_rows() -> dict[str, list[list[object]]]:
    rows = deepcopy(minimal_workbook_rows())

    rows["Valve_Plan"][0] = [
        "Valve_ID",
        "Order_ID",
        "Customer",
        "Dispatch_Date",
        "Assembly_Date",
        "Value_Cr",
        "Priority",
    ]
    rows["Valve_Plan"][1] = ["V-100", "O-100", "Acme", "2026-05-01", "2026-04-22", 1.25, "A"]
    rows["Valve_Plan"].append(["V-200", "O-200", "Beta", "2026-05-02", "2026-04-24", 0.5, "B"])

    rows["Component_Status"][0] = [
        "Valve_ID",
        "Component_Line_No",
        "Component",
        "Qty",
        "Fabrication_Required",
        "Fabrication_Complete",
        "Expected_Ready_Date",
        "Critical",
        "Ready_Date_Type",
    ]
    rows["Component_Status"][1] = ["V-100", 1, "Body", 1, "N", "Y", "2026-04-21", "Y", "CONFIRMED"]
    rows["Component_Status"].append(["V-200", 1, "Bonnet", 1, "N", "Y", "2026-04-21", "Y", "CONFIRMED"])

    rows["Routing_Master"][0] = [
        "Component",
        "Operation_No",
        "Operation_Name",
        "Machine_Type",
        "Std_Total_Hrs",
        "Subcontract_Allowed",
        "Vendor_Process",
    ]
    rows["Routing_Master"][1] = ["Body", 10, "HBM roughing", "HBM", 8, "Y", "HBM"]
    rows["Routing_Master"].append(["Body", 20, "VTL finish", "VTL", 4, "N", ""])
    rows["Routing_Master"].append(["Bonnet", 10, "HBM finish", "HBM", 8, "Y", "HBM"])

    rows["Machine_Master"][0] = [
        "Machine_ID",
        "Machine_Type",
        "Hours_per_Day",
        "Efficiency_Percent",
        "Buffer_Days",
        "Active",
    ]
    rows["Machine_Master"][1] = ["HBM-1", "HBM", 8, 100, 1, "Y"]
    rows["Machine_Master"].append(["VTL-1", "VTL", 8, 100, 3, "Y"])

    rows["Vendor_Master"][0] = [
        "Vendor_ID",
        "Vendor_Name",
        "Primary_Process",
        "Turnaround_Days",
        "Transport_Days_Total",
        "Capacity_Rating",
        "Reliability",
        "Approved",
    ]
    rows["Vendor_Master"][1] = ["VEN-1", "Vendor One", "HBM", 0, 0, "Medium", "A", "Y"]
    rows["Vendor_Master"].append(["VEN-2", "Vendor Two", "VTL", 2, 1, "Low", "B", "Y"])

    return rows
