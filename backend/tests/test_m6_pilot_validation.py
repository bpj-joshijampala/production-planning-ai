from collections.abc import Generator
from datetime import date, timedelta
from io import BytesIO
from pathlib import Path

import pytest
from alembic import command
from alembic.config import Config
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.core.config import get_settings
from app.main import create_app


FIXTURE_PATH = Path(__file__).parent / "fixtures" / "machine_shop_sample_input.xlsx"
XLSX_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PLANNING_START_DATE = "2026-04-21"
PLANNING_HORIZON_DAYS = 7
PILOT_MANAGEMENT_METRICS = {
    "throughput_gap_cr": 0.75,
    "overloaded_machines": 1,
    "subcontract_recommendations": 3,
    "assembly_risk_valves": 1,
    "flow_blockers": 4,
}
V1_EXPORT_TYPES = (
    "MACHINE_LOAD",
    "SUBCONTRACT_PLAN",
    "VALVE_READINESS",
    "FLOW_BLOCKER",
    "DAILY_EXECUTION",
    "WEEKLY_PLANNING",
    "A3_PLANNING",
)


@pytest.fixture(name="client")
def fixture_client(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> Generator[TestClient, None, None]:
    database_path = tmp_path / "m6_pilot_validation.sqlite3"
    upload_dir = tmp_path / "uploads"
    export_dir = tmp_path / "exports"

    monkeypatch.setenv("DATABASE_URL", f"sqlite:///{database_path.as_posix()}")
    monkeypatch.setenv("UPLOAD_DIR", upload_dir.as_posix())
    monkeypatch.setenv("EXPORT_DIR", export_dir.as_posix())
    get_settings.cache_clear()

    command.upgrade(Config("alembic.ini"), "head")

    with TestClient(create_app()) as test_client:
        yield test_client

    get_settings.cache_clear()


def test_pilot_workflow_validates_planner_hod_management_and_export_paths(client: TestClient) -> None:
    upload_payload = _upload_golden_workbook(client)
    planning_run_id = _create_and_calculate_pilot_run(client=client, upload_batch_id=upload_payload["id"])

    dashboard_payload = _get_json(client, f"/api/v1/planning-runs/{planning_run_id}/dashboard")
    assert {
        metric: dashboard_payload[metric]
        for metric in PILOT_MANAGEMENT_METRICS
    } == PILOT_MANAGEMENT_METRICS
    assert dashboard_payload["active_valves"] == 3
    assert dashboard_payload["planned_throughput_value_cr"] == 1.75

    machine_load_payload = _get_json(
        client,
        f"/api/v1/planning-runs/{planning_run_id}/machine-load?page=1&page_size=100&sort=machine_type",
    )
    assert machine_load_payload["total"] == 2
    _assert_machine_load_matches_manual_workbook_calculation(machine_load_payload["items"])

    queue_payload = _get_json(
        client,
        f"/api/v1/planning-runs/{planning_run_id}/machine-load/HBM/queue?page=1&page_size=100",
    )
    assert queue_payload["total"] == 3
    assert [row["operation_name"] for row in queue_payload["items"]] == [
        "HBM roughing",
        "HBM seat prep",
        "HBM finish",
    ]

    recommendations_payload = _get_json(
        client,
        f"/api/v1/planning-runs/{planning_run_id}/subcontract-recommendations?page=1&page_size=100",
    )
    assert recommendations_payload["total"] == 4
    assert all(row["explanation"] for row in recommendations_payload["items"])
    recommendation = next(
        row
        for row in recommendations_payload["items"]
        if row["recommendation_type"] == "BATCH_SUBCONTRACT_OPPORTUNITY"
    )

    override_response = client.post(
        "/api/v1/planner-overrides",
        json={
            "planning_run_id": planning_run_id,
            "entity_type": "RECOMMENDATION",
            "entity_id": recommendation["id"],
            "override_decision": "ACCEPT",
            "reason": "Pilot planner accepts the HBM batch subcontract recommendation.",
            "remarks": "Reviewed during M6-E5 pilot validation.",
        },
    )
    assert override_response.status_code == 201
    override_payload = override_response.json()
    assert override_payload["recommendation_id"] == recommendation["id"]
    assert override_payload["original_recommendation"] == "BATCH_SUBCONTRACT_OPPORTUNITY"
    assert override_payload["override_decision"] == "ACCEPT"
    assert override_payload["user_display_name"] == "Development Planner"
    assert override_payload["stale_flag"] is False

    accepted_recommendations_payload = _get_json(
        client,
        f"/api/v1/planning-runs/{planning_run_id}/subcontract-recommendations?page=1&page_size=100&status=ACCEPTED",
    )
    assert accepted_recommendations_payload["total"] == 1
    assert accepted_recommendations_payload["items"][0]["id"] == recommendation["id"]

    action_log_payload = _get_json(client, f"/api/v1/planning-runs/{planning_run_id}/planner-overrides")
    assert action_log_payload["current_override_count"] == 1
    assert action_log_payload["stale_override_count"] == 0
    assert action_log_payload["overrides"][0]["reason"] == (
        "Pilot planner accepts the HBM batch subcontract recommendation."
    )

    exported_sheet_names = _generate_and_download_v1_exports(
        client=client,
        planning_run_id=planning_run_id,
        accepted_recommendation=recommendation,
        override_reason="Pilot planner accepts the HBM batch subcontract recommendation.",
        override_remarks="Reviewed during M6-E5 pilot validation.",
    )
    assert exported_sheet_names["WEEKLY_PLANNING"] == [
        "Export_Info",
        "Weekly_Summary",
        "Machine_Load",
        "Valve_Readiness",
        "Flow_Blockers",
        "Subcontract_Plan",
    ]
    assert exported_sheet_names["A3_PLANNING"] == ["Export_Info", "A3_Planning"]

    export_history_payload = _get_json(
        client,
        f"/api/v1/planning-runs/{planning_run_id}/exports?latest_only=true&page=1&page_size=100",
    )
    assert export_history_payload["total"] == len(V1_EXPORT_TYPES)
    assert {row["report_type"] for row in export_history_payload["items"]} == set(V1_EXPORT_TYPES)


def _upload_golden_workbook(client: TestClient) -> dict[str, object]:
    response = client.post(
        "/api/v1/uploads",
        files={"file": (FIXTURE_PATH.name, FIXTURE_PATH.read_bytes(), XLSX_MIME_TYPE)},
    )
    assert response.status_code == 201
    payload = response.json()
    assert payload["status"] == "VALIDATED"
    assert payload["validation_error_count"] == 0
    assert payload["validation_warning_count"] == 0

    validation_payload = _get_json(client, f"/api/v1/uploads/{payload['id']}/validation-issues")
    assert validation_payload == {
        "upload_batch_id": payload["id"],
        "summary": {"blocking": 0, "warning": 0, "total": 0},
        "issues": [],
    }
    return payload


def _create_and_calculate_pilot_run(*, client: TestClient, upload_batch_id: str) -> str:
    create_response = client.post(
        "/api/v1/planning-runs",
        json={
            "upload_batch_id": upload_batch_id,
            "planning_start_date": PLANNING_START_DATE,
            "planning_horizon_days": PLANNING_HORIZON_DAYS,
        },
    )
    assert create_response.status_code == 201
    planning_run_id = create_response.json()["id"]

    calculate_response = client.post(f"/api/v1/planning-runs/{planning_run_id}/calculate")
    assert calculate_response.status_code == 200
    assert calculate_response.json()["status"] == "CALCULATED"
    assert calculate_response.json()["error_message"] is None
    return planning_run_id


def _assert_machine_load_matches_manual_workbook_calculation(api_rows: list[dict[str, object]]) -> None:
    manual_rows = _manual_machine_load_from_workbook(
        workbook_path=FIXTURE_PATH,
        planning_start_date=date.fromisoformat(PLANNING_START_DATE),
        planning_horizon_days=PLANNING_HORIZON_DAYS,
    )
    actual_by_machine = {str(row["machine_type"]): row for row in api_rows}

    assert set(actual_by_machine) == set(manual_rows)
    for machine_type, expected in manual_rows.items():
        actual = actual_by_machine[machine_type]
        assert actual["total_operation_hours"] == pytest.approx(expected["total_operation_hours"])
        assert actual["capacity_hours_per_day"] == pytest.approx(expected["capacity_hours_per_day"])
        assert actual["load_days"] == pytest.approx(expected["load_days"])
        assert actual["buffer_days"] == pytest.approx(expected["buffer_days"])
        assert actual["overload_flag"] is expected["overload_flag"]
        assert actual["overload_days"] == pytest.approx(expected["overload_days"])
        assert actual["spare_capacity_days"] == pytest.approx(expected["spare_capacity_days"])
        assert actual["underutilized_flag"] is expected["underutilized_flag"]


def _manual_machine_load_from_workbook(
    *,
    workbook_path: Path,
    planning_start_date: date,
    planning_horizon_days: int,
) -> dict[str, dict[str, float | bool]]:
    planning_end_date = planning_start_date + timedelta(days=planning_horizon_days)
    workbook = load_workbook(workbook_path, data_only=True)
    components = _sheet_records(workbook, "Component_Status")
    routing_by_component: dict[str, list[dict[str, object]]] = {}
    for row in _sheet_records(workbook, "Routing_Master"):
        routing_by_component.setdefault(str(row["Component"]), []).append(row)

    load_hours_by_machine: dict[str, float] = {}
    for component in components:
        expected_ready_date = date.fromisoformat(str(component["Expected_Ready_Date"]))
        if not _is_yes(component["Fabrication_Complete"]) or expected_ready_date > planning_end_date:
            continue

        qty = float(component["Qty"])
        for operation in routing_by_component.get(str(component["Component"]), []):
            machine_type = str(operation["Machine_Type"])
            operation_hours = qty * _operation_standard_hours(operation)
            load_hours_by_machine[machine_type] = load_hours_by_machine.get(machine_type, 0.0) + operation_hours

    manual_rows: dict[str, dict[str, float | bool]] = {}
    for machine in _sheet_records(workbook, "Machine_Master"):
        machine_type = str(machine["Machine_Type"])
        if not _is_yes(machine["Active"]):
            continue

        capacity_hours_per_day = float(machine["Hours_per_Day"]) * float(machine["Efficiency_Percent"]) / 100
        total_operation_hours = load_hours_by_machine.get(machine_type, 0.0)
        load_days = 0.0 if capacity_hours_per_day == 0 else total_operation_hours / capacity_hours_per_day
        buffer_days = float(machine["Buffer_Days"])
        overload_days = max(load_days - buffer_days, 0.0)
        spare_capacity_days = max(buffer_days - load_days, 0.0)
        underutilized_threshold_days = 0.5 * buffer_days
        manual_rows[machine_type] = {
            "total_operation_hours": total_operation_hours,
            "capacity_hours_per_day": capacity_hours_per_day,
            "load_days": load_days,
            "buffer_days": buffer_days,
            "overload_flag": overload_days > 0,
            "overload_days": overload_days,
            "spare_capacity_days": spare_capacity_days,
            "underutilized_flag": load_days < underutilized_threshold_days,
        }

    return manual_rows


def _operation_standard_hours(operation: dict[str, object]) -> float:
    std_total_hours = operation.get("Std_Total_Hrs")
    if std_total_hours is not None:
        return float(std_total_hours)
    return float(operation.get("Std_Setup_Hrs") or 0.0) + float(operation.get("Std_Run_Hrs") or 0.0)


def _generate_and_download_v1_exports(
    *,
    client: TestClient,
    planning_run_id: str,
    accepted_recommendation: dict[str, object],
    override_reason: str,
    override_remarks: str,
) -> dict[str, list[str]]:
    sheet_names_by_report_type: dict[str, list[str]] = {}
    for report_type in V1_EXPORT_TYPES:
        export_response = client.post(
            f"/api/v1/planning-runs/{planning_run_id}/exports",
            json={"report_type": report_type, "file_format": "XLSX"},
        )
        assert export_response.status_code == 201
        export_payload = export_response.json()
        assert export_payload["generated_by_user_display_name"] == "Development Planner"
        assert export_payload["metadata"]["sheet_names"]

        download_response = client.get(export_payload["download_url"])
        assert download_response.status_code == 200
        assert download_response.headers["x-report-export-id"] == export_payload["id"]

        workbook = load_workbook(BytesIO(download_response.content), data_only=True)
        assert workbook["Export_Info"]["A1"].value == "Field"
        if report_type == "SUBCONTRACT_PLAN":
            _assert_subcontract_export_contains_accepted_decision(
                workbook=workbook,
                accepted_recommendation=accepted_recommendation,
            )
        if report_type == "A3_PLANNING":
            _assert_a3_export_contains_planner_override(
                workbook=workbook,
                accepted_recommendation=accepted_recommendation,
                override_reason=override_reason,
                override_remarks=override_remarks,
            )
        sheet_names_by_report_type[report_type] = workbook.sheetnames

    return sheet_names_by_report_type


def _assert_subcontract_export_contains_accepted_decision(
    *,
    workbook,
    accepted_recommendation: dict[str, object],
) -> None:  # type: ignore[no-untyped-def]
    rows = _sheet_records(workbook, "Subcontract_Plan")
    accepted_row = next(
        row
        for row in rows
        if row["Valve_ID"] == accepted_recommendation["valve_id"]
        and row["Component_Line_No"] == accepted_recommendation["component_line_no"]
        and row["Operation_Name"] == accepted_recommendation["operation_name"]
    )

    assert accepted_row["Recommendation_Type"] == "BATCH_SUBCONTRACT_OPPORTUNITY"
    assert accepted_row["Suggested_Vendor_ID"] == accepted_recommendation["suggested_vendor_id"]
    assert accepted_row["Status"] == "ACCEPTED"


def _assert_a3_export_contains_planner_override(
    *,
    workbook,
    accepted_recommendation: dict[str, object],
    override_reason: str,
    override_remarks: str,
) -> None:  # type: ignore[no-untyped-def]
    rows = _sheet_records(workbook, "A3_Planning")
    override_row = next(
        row
        for row in rows
        if row["Section"] == "Planner overrides"
        and row["Item"] == accepted_recommendation["id"]
    )

    assert override_row["Value"] == "ACCEPT"
    assert override_reason in str(override_row["Detail"])
    assert override_remarks == override_row["Recommended_Action"]


def _get_json(client: TestClient, url: str) -> dict[str, object]:
    response = client.get(url)
    assert response.status_code == 200
    return response.json()


def _sheet_records(workbook, sheet_name: str) -> list[dict[str, object]]:  # type: ignore[no-untyped-def]
    worksheet = workbook[sheet_name]
    headers = [str(cell.value) for cell in worksheet[1]]
    return [
        dict(zip(headers, row_values, strict=True))
        for row_values in worksheet.iter_rows(min_row=2, values_only=True)
    ]


def _is_yes(value: object) -> bool:
    return str(value).strip().upper() in {"Y", "YES", "TRUE", "1"}
