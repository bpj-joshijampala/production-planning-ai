from collections.abc import Generator
from io import BytesIO
from pathlib import Path
import json

from alembic import command
from alembic.config import Config
from fastapi.testclient import TestClient
from openpyxl import load_workbook
import pytest
from sqlalchemy import func, select

from app.core.config import get_settings
from app.db.session import create_session_factory
from app.main import create_app
from app.models.canonical import ComponentStatus, Machine, RoutingOperation, Valve, Vendor
from app.models.output import (
    FlowBlocker,
    IncomingLoadItem,
    MachineLoadSummary,
    PlannedOperation,
    Recommendation,
    ReportExport,
    ThroughputSummary,
    ValveReadinessSummary,
    VendorLoadSummary,
)
from app.models.planning_run import PlanningRun, PlanningSnapshot
from app.models.upload import ImportStagingRow, ImportValidationIssue, UploadBatch


FIXTURE_PATH = Path(__file__).parent / "fixtures" / "machine_shop_sample_input.xlsx"
XLSX_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

EXPECTED_SHEET_HEADERS = {
    "Valve_Plan": [
        "Valve_ID",
        "Order_ID",
        "Customer",
        "Valve_Type",
        "Dispatch_Date",
        "Assembly_Date",
        "Value_Cr",
        "Priority",
        "Status",
        "Remarks",
    ],
    "Component_Status": [
        "Valve_ID",
        "Component_Line_No",
        "Component",
        "Qty",
        "Fabrication_Required",
        "Fabrication_Complete",
        "Expected_Ready_Date",
        "Critical",
        "Ready_Date_Type",
        "Current_Location",
        "Comments",
    ],
    "Routing_Master": [
        "Component",
        "Operation_No",
        "Operation_Name",
        "Machine_Type",
        "Alt_Machine",
        "Std_Setup_Hrs",
        "Std_Run_Hrs",
        "Std_Total_Hrs",
        "Subcontract_Allowed",
        "Vendor_Process",
        "Notes",
    ],
    "Machine_Master": [
        "Machine_ID",
        "Machine_Type",
        "Description",
        "Hours_per_Day",
        "Efficiency_Percent",
        "Buffer_Days",
        "Active",
    ],
    "Vendor_Master": [
        "Vendor_ID",
        "Vendor_Name",
        "Primary_Process",
        "Turnaround_Days",
        "Transport_Days_Total",
        "Capacity_Rating",
        "Reliability",
        "Approved",
        "Comments",
    ],
}

EXPECTED_STAGING_COUNTS = {
    "Valve_Plan": 3,
    "Component_Status": 4,
    "Routing_Master": 5,
    "Machine_Master": 2,
    "Vendor_Master": 2,
}

EXPECTED_CANONICAL_COUNTS = {
    "valves": 3,
    "component_statuses": 4,
    "routing_operations": 5,
    "machines": 2,
    "vendors": 2,
}

EXPORT_EXPECTATIONS = {
    "MACHINE_LOAD": (
        "Machine_Load",
        [
            "Machine_Type",
            "Total_Operation_Hours",
            "Capacity_Hours_Per_Day",
            "Load_Days",
            "Buffer_Days",
            "Overload_Flag",
            "Overload_Days",
            "Spare_Capacity_Days",
            "Underutilized_Flag",
            "Batch_Risk_Flag",
            "Status",
        ],
        2,
    ),
    "SUBCONTRACT_PLAN": (
        "Subcontract_Plan",
        [
            "Recommendation_Type",
            "Valve_ID",
            "Component_Line_No",
            "Component",
            "Operation_Name",
            "Machine_Type",
            "Suggested_Vendor_ID",
            "Suggested_Vendor_Name",
            "Internal_Wait_Days",
            "Internal_Completion_Days",
            "Vendor_Total_Days",
            "Vendor_Gain_Days",
            "Batch_Candidate_Count",
            "Batch_Opportunity",
            "Status",
            "Explanation",
        ],
        3,
    ),
    "VALVE_READINESS": (
        "Valve_Readiness",
        [
            "Valve_ID",
            "Customer",
            "Assembly_Date",
            "Dispatch_Date",
            "Value_Cr",
            "Total_Components",
            "Ready_Components",
            "Required_Components",
            "Ready_Required_Count",
            "Pending_Required_Count",
            "Full_Kit",
            "Near_Ready",
            "Expected_Completion_Date",
            "Assembly_Delay_Days",
            "Status",
            "Risk_Reason",
            "Valve_Flow_Gap_Days",
            "Valve_Flow_Imbalance",
        ],
        3,
    ),
    "FLOW_BLOCKER": (
        "Flow_Blockers",
        [
            "Severity",
            "Blocker_Type",
            "Valve_ID",
            "Component_Line_No",
            "Component",
            "Operation_Name",
            "Cause",
            "Recommended_Action",
        ],
        4,
    ),
    "DAILY_EXECUTION": (
        "Daily_Execution",
        [
            "Date",
            "Machine_Type",
            "Queue_Sequence",
            "Valve_ID",
            "Component_Line_No",
            "Component",
            "Operation_Name",
            "Planned_Action",
            "Internal_Wait_Days",
            "Internal_Completion_Date",
            "Extreme_Delay_Flag",
        ],
        4,
    ),
}


@pytest.fixture(name="client")
def fixture_client(tmp_path, monkeypatch) -> Generator[TestClient, None, None]:  # type: ignore[no-untyped-def]
    database_path = tmp_path / "m5_golden_workbook.sqlite3"
    upload_dir = tmp_path / "uploads"
    export_dir = tmp_path / "exports"

    monkeypatch.setenv("DATABASE_URL", f"sqlite:///{database_path.as_posix()}")
    monkeypatch.setenv("UPLOAD_DIR", upload_dir.as_posix())
    monkeypatch.setenv("EXPORT_DIR", export_dir.as_posix())
    get_settings.cache_clear()

    command.upgrade(Config("alembic.ini"), "head")

    with TestClient(create_app()) as test_client:
        yield test_client

    get_settings.cache_clear()


def test_machine_shop_sample_workbook_imports_calculates_and_exports(client: TestClient) -> None:
    workbook_content = _read_and_assert_golden_workbook_structure()

    upload_response = client.post(
        "/api/v1/uploads",
        files={"file": (FIXTURE_PATH.name, workbook_content, XLSX_MIME_TYPE)},
    )

    assert upload_response.status_code == 201
    upload_payload = upload_response.json()
    assert upload_payload["original_filename"] == "machine_shop_sample_input.xlsx"
    assert upload_payload["status"] == "VALIDATED"
    assert upload_payload["validation_error_count"] == 0
    assert upload_payload["validation_warning_count"] == 0

    validation_response = client.get(f"/api/v1/uploads/{upload_payload['id']}/validation-issues")
    assert validation_response.status_code == 200
    assert validation_response.json()["summary"] == {"blocking": 0, "warning": 0, "total": 0}
    assert validation_response.json()["issues"] == []

    session_factory = create_session_factory()
    with session_factory() as session:
        assert _staging_counts(session=session, upload_batch_id=upload_payload["id"]) == EXPECTED_STAGING_COUNTS
        assert (
            session.scalar(
                select(func.count())
                .select_from(ImportValidationIssue)
                .where(ImportValidationIssue.upload_batch_id == upload_payload["id"])
            )
            == 0
        )

    planning_run_response = client.post(
        "/api/v1/planning-runs",
        json={
            "upload_batch_id": upload_payload["id"],
            "planning_start_date": "2026-04-21",
            "planning_horizon_days": 7,
        },
    )

    assert planning_run_response.status_code == 201
    planning_run_payload = planning_run_response.json()
    planning_run_id = planning_run_payload["id"]
    assert planning_run_payload["status"] == "CREATED"
    assert planning_run_payload["snapshot_id"] is not None
    assert planning_run_payload["canonical_counts"] == EXPECTED_CANONICAL_COUNTS

    calculate_response = client.post(f"/api/v1/planning-runs/{planning_run_id}/calculate")

    assert calculate_response.status_code == 200
    assert calculate_response.json()["status"] == "CALCULATED"
    assert calculate_response.json()["error_message"] is None
    assert calculate_response.json()["canonical_counts"] == EXPECTED_CANONICAL_COUNTS

    _assert_golden_database_outputs(planning_run_id=planning_run_id, upload_batch_id=upload_payload["id"])
    _assert_golden_dashboard_outputs(client=client, planning_run_id=planning_run_id)
    _assert_golden_report_exports(client=client, planning_run_id=planning_run_id)


def _read_and_assert_golden_workbook_structure() -> bytes:
    assert FIXTURE_PATH.exists()
    workbook_content = FIXTURE_PATH.read_bytes()
    workbook = load_workbook(BytesIO(workbook_content), data_only=True)

    assert workbook.sheetnames == list(EXPECTED_SHEET_HEADERS)
    for sheet_name, expected_headers in EXPECTED_SHEET_HEADERS.items():
        worksheet = workbook[sheet_name]
        assert [cell.value for cell in worksheet[1]] == expected_headers
        assert worksheet.max_row - 1 == EXPECTED_STAGING_COUNTS[sheet_name]

    return workbook_content


def _assert_golden_database_outputs(*, planning_run_id: str, upload_batch_id: str) -> None:
    session_factory = create_session_factory()
    with session_factory() as session:
        upload = session.get(UploadBatch, upload_batch_id)
        planning_run = session.get(PlanningRun, planning_run_id)
        snapshot = session.scalar(
            select(PlanningSnapshot)
            .where(PlanningSnapshot.planning_run_id == planning_run_id)
            .order_by(PlanningSnapshot.created_at.desc(), PlanningSnapshot.id.desc())
            .limit(1)
        )

        assert upload is not None
        assert upload.status == "CALCULATED"
        assert planning_run is not None
        assert planning_run.status == "CALCULATED"
        assert planning_run.calculated_at is not None
        assert snapshot is not None
        assert '"planning_horizon_days":7' in snapshot.snapshot_json

        assert _model_count(session=session, model=Valve, planning_run_id=planning_run_id) == 3
        assert _model_count(session=session, model=ComponentStatus, planning_run_id=planning_run_id) == 4
        assert _model_count(session=session, model=RoutingOperation, planning_run_id=planning_run_id) == 5
        assert _model_count(session=session, model=Machine, planning_run_id=planning_run_id) == 2
        assert _model_count(session=session, model=Vendor, planning_run_id=planning_run_id) == 2
        assert _model_count(session=session, model=IncomingLoadItem, planning_run_id=planning_run_id) == 3
        assert _model_count(session=session, model=PlannedOperation, planning_run_id=planning_run_id) == 4
        assert _model_count(session=session, model=MachineLoadSummary, planning_run_id=planning_run_id) == 2
        assert _model_count(session=session, model=ValveReadinessSummary, planning_run_id=planning_run_id) == 3
        assert _model_count(session=session, model=FlowBlocker, planning_run_id=planning_run_id) == 4
        assert _model_count(session=session, model=Recommendation, planning_run_id=planning_run_id) == 4
        assert _model_count(session=session, model=ThroughputSummary, planning_run_id=planning_run_id) == 1
        assert _model_count(session=session, model=VendorLoadSummary, planning_run_id=planning_run_id) == 2

        incoming_load = list(
            session.scalars(
                select(IncomingLoadItem)
                .where(IncomingLoadItem.planning_run_id == planning_run_id)
                .order_by(IncomingLoadItem.sort_sequence.asc())
            )
        )
        assert [
            (
                row.valve_id,
                row.component_line_no,
                row.component,
                row.availability_date,
                json.loads(row.machine_types_json or "[]"),
                row.sort_sequence,
            )
            for row in incoming_load
        ] == [
            ("V-100", 1, "Body", "2026-04-21", ["HBM", "VTL"], 1),
            ("V-100", 2, "Seat", "2026-04-21", ["HBM"], 2),
            ("V-200", 1, "Bonnet", "2026-04-21", ["HBM"], 3),
        ]

        planned_operations = list(
            session.scalars(
                select(PlannedOperation)
                .where(PlannedOperation.planning_run_id == planning_run_id)
                .order_by(PlannedOperation.sort_sequence.asc(), PlannedOperation.operation_no.asc())
            )
        )
        assert [
            (
                row.valve_id,
                row.component_line_no,
                row.component,
                row.operation_no,
                row.operation_name,
                row.machine_type,
                row.internal_wait_days,
                row.processing_time_days,
                row.internal_completion_offset_days,
                row.recommendation_status,
                row.extreme_delay_flag,
            )
            for row in planned_operations
        ] == [
            ("V-100", 1, "Body", 10, "HBM roughing", "HBM", 0.0, 1.0, 1.0, "SUBCONTRACT", 0),
            ("V-100", 1, "Body", 20, "VTL finish", "VTL", 0.0, 0.5, 1.5, "OK_INTERNAL", 0),
            ("V-100", 2, "Seat", 10, "HBM seat prep", "HBM", 1.0, 1.0, 2.0, "SUBCONTRACT", 0),
            ("V-200", 1, "Bonnet", 10, "HBM finish", "HBM", 2.0, 1.0, 3.0, "SUBCONTRACT", 0),
        ]

        machine_load = list(
            session.scalars(
                select(MachineLoadSummary)
                .where(MachineLoadSummary.planning_run_id == planning_run_id)
                .order_by(MachineLoadSummary.machine_type.asc())
            )
        )
        assert [
            (
                row.machine_type,
                row.total_operation_hours,
                row.load_days,
                row.buffer_days,
                row.overload_flag,
                row.overload_days,
                row.spare_capacity_days,
                row.underutilized_flag,
                row.batch_risk_flag,
                row.status,
            )
            for row in machine_load
        ] == [
            ("HBM", 24.0, 3.0, 1.0, 1, 2.0, 0.0, 0, 1, "OVERLOADED"),
            ("VTL", 4.0, 0.5, 3.0, 0, 0.0, 2.5, 1, 0, "UNDERUTILIZED"),
        ]

        readiness = list(
            session.scalars(
                select(ValveReadinessSummary)
                .where(ValveReadinessSummary.planning_run_id == planning_run_id)
                .order_by(ValveReadinessSummary.valve_id.asc())
            )
        )
        assert [
            (
                row.valve_id,
                row.ready_components,
                row.required_components,
                row.pending_required_count,
                row.full_kit_flag,
                row.near_ready_flag,
                row.valve_expected_completion_date,
                row.otd_delay_days,
                row.otd_risk_flag,
                row.readiness_status,
                row.risk_reason,
                row.valve_flow_gap_days,
            )
            for row in readiness
        ] == [
            ("V-100", 2, 2, 0, 1, 0, "2026-04-23", 1.0, 1, "AT_RISK", "Assembly delay", 0.5),
            ("V-200", 1, 1, 0, 1, 0, "2026-04-24", 0.0, 0, "READY", None, 0.0),
            ("V-300", 0, 1, 1, 0, 1, None, 0.0, 0, "DATA_INCOMPLETE", "Data issue", None),
        ]

        recommendations = list(
            session.scalars(
                select(Recommendation)
                .where(Recommendation.planning_run_id == planning_run_id)
                .order_by(
                    Recommendation.valve_id.asc(),
                    Recommendation.component_line_no.asc(),
                    Recommendation.operation_name.asc(),
                )
            )
        )
        assert [
            (
                row.valve_id,
                row.component_line_no,
                row.component,
                row.operation_name,
                row.machine_type,
                row.recommendation_type,
                row.suggested_vendor_id,
                row.subcontract_batch_candidate_count,
                row.batch_subcontract_opportunity_flag,
                json.loads(row.reason_codes_json),
            )
            for row in recommendations
        ] == [
            (
                "V-100",
                1,
                "Body",
                "HBM roughing",
                "HBM",
                "BATCH_SUBCONTRACT_OPPORTUNITY",
                "VEN-1",
                3,
                1,
                ["PRIMARY_OVERLOADED", "SUBCONTRACT_FEASIBLE", "BATCH_SUBCONTRACT_OPPORTUNITY"],
            ),
            ("V-100", 1, "Body", "VTL finish", "VTL", "OK_INTERNAL", None, None, 0, ["OK_INTERNAL"]),
            (
                "V-100",
                2,
                "Seat",
                "HBM seat prep",
                "HBM",
                "BATCH_SUBCONTRACT_OPPORTUNITY",
                "VEN-1",
                3,
                1,
                ["PRIMARY_OVERLOADED", "SUBCONTRACT_FEASIBLE", "BATCH_SUBCONTRACT_OPPORTUNITY"],
            ),
            (
                "V-200",
                1,
                "Bonnet",
                "HBM finish",
                "HBM",
                "BATCH_SUBCONTRACT_OPPORTUNITY",
                "VEN-1",
                3,
                1,
                ["PRIMARY_OVERLOADED", "SUBCONTRACT_FEASIBLE", "BATCH_SUBCONTRACT_OPPORTUNITY"],
            ),
        ]

        blockers = list(
            session.scalars(
                select(FlowBlocker)
                .where(FlowBlocker.planning_run_id == planning_run_id)
                .order_by(FlowBlocker.blocker_type.asc(), FlowBlocker.valve_id.asc())
            )
        )
        assert [
            (row.blocker_type, row.valve_id, row.component_line_no, row.component, row.severity)
            for row in blockers
        ] == [
            ("BATCH_RISK", None, None, None, "INFO"),
            ("MACHINE_OVERLOAD", None, None, None, "WARNING"),
            ("MISSING_COMPONENT", "V-300", 1, "Disc", "WARNING"),
            ("VENDOR_OVERLOADED", None, None, None, "WARNING"),
        ]

        throughput = session.scalar(
            select(ThroughputSummary).where(ThroughputSummary.planning_run_id == planning_run_id)
        )
        assert throughput is not None
        assert throughput.target_throughput_value_cr == pytest.approx(2.5)
        assert throughput.planned_throughput_value_cr == pytest.approx(1.75)
        assert throughput.throughput_gap_cr == pytest.approx(0.75)
        assert throughput.throughput_risk_flag == 1

        vendor_load = list(
            session.scalars(
                select(VendorLoadSummary)
                .where(VendorLoadSummary.planning_run_id == planning_run_id)
                .order_by(VendorLoadSummary.vendor_id.asc())
            )
        )
        assert [
            (
                row.vendor_id,
                row.primary_process,
                row.vendor_recommended_jobs,
                row.max_recommended_jobs_per_horizon,
                row.selected_vendor_overloaded_flag,
                row.status,
            )
            for row in vendor_load
        ] == [
            ("VEN-1", "HBM", 3, 3, 1, "VENDOR_OVERLOADED"),
            ("VEN-2", "VTL", 0, 1, 0, "OK"),
        ]


def _assert_golden_dashboard_outputs(*, client: TestClient, planning_run_id: str) -> None:
    dashboard_response = client.get(f"/api/v1/planning-runs/{planning_run_id}/dashboard")
    assert dashboard_response.status_code == 200
    assert dashboard_response.json() == {
        "planning_run_id": planning_run_id,
        "active_valves": 3,
        "active_value_cr": 2.5,
        "planned_throughput_value_cr": 1.75,
        "throughput_gap_cr": 0.75,
        "overloaded_machines": 1,
        "underutilized_machines": 1,
        "flow_blockers": 4,
        "assembly_risk_valves": 1,
        "subcontract_recommendations": 3,
        "batch_risks": 1,
    }

    throughput_response = client.get(f"/api/v1/planning-runs/{planning_run_id}/throughput")
    assert throughput_response.status_code == 200
    assert throughput_response.json() == {
        "planning_run_id": planning_run_id,
        "target_throughput_value_cr": 2.5,
        "planned_throughput_value_cr": 1.75,
        "throughput_gap_cr": 0.75,
        "throughput_risk_flag": True,
    }

    incoming_load_response = client.get(f"/api/v1/planning-runs/{planning_run_id}/incoming-load?page=1&page_size=100")
    assert incoming_load_response.status_code == 200
    assert incoming_load_response.json()["total"] == 3
    assert incoming_load_response.json()["items"][0]["valve_id"] == "V-100"
    assert incoming_load_response.json()["items"][0]["machine_types"] == ["HBM", "VTL"]

    machine_load_response = client.get(f"/api/v1/planning-runs/{planning_run_id}/machine-load?page=1&page_size=100")
    assert machine_load_response.status_code == 200
    assert [(row["machine_type"], row["status"]) for row in machine_load_response.json()["items"]] == [
        ("HBM", "OVERLOADED"),
        ("VTL", "UNDERUTILIZED"),
    ]

    recommendations_response = client.get(
        f"/api/v1/planning-runs/{planning_run_id}/subcontract-recommendations?page=1&page_size=100"
    )
    assert recommendations_response.status_code == 200
    assert recommendations_response.json()["total"] == 4
    assert {row["recommendation_type"] for row in recommendations_response.json()["items"]} == {
        "BATCH_SUBCONTRACT_OPPORTUNITY",
        "OK_INTERNAL",
    }

    blockers_response = client.get(f"/api/v1/planning-runs/{planning_run_id}/flow-blockers?page=1&page_size=100")
    assert blockers_response.status_code == 200
    assert blockers_response.json()["total"] == 4
    assert {row["blocker_type"] for row in blockers_response.json()["items"]} == {
        "MACHINE_OVERLOAD",
        "BATCH_RISK",
        "MISSING_COMPONENT",
        "VENDOR_OVERLOADED",
    }

    component_response = client.get(
        f"/api/v1/planning-runs/{planning_run_id}/component-status?valve_id=V-300&page=1&page_size=100"
    )
    assert component_response.status_code == 200
    assert component_response.json()["items"] == [
        {
            "valve_id": "V-300",
            "customer": "Gamma Energy",
            "component_line_no": 1,
            "component": "Disc",
            "current_location": "Fabrication",
            "fabrication_complete": False,
            "critical": True,
            "availability_date": "2026-04-30",
            "date_confidence": "EXPECTED",
            "next_operation_name": None,
            "next_machine_type": None,
            "internal_wait_days": None,
            "status": "BLOCKED",
            "blocker_types": ["MISSING_COMPONENT"],
            "blocker_summary": "Required component Disc availability_date 2026-04-30 is outside planning horizon ending 2026-04-28.",
        }
    ]

    assembly_risk_response = client.get(f"/api/v1/planning-runs/{planning_run_id}/assembly-risk?page=1&page_size=100")
    assert assembly_risk_response.status_code == 200
    assert assembly_risk_response.json()["total"] == 1
    assert assembly_risk_response.json()["items"][0]["valve_id"] == "V-100"


def _assert_golden_report_exports(*, client: TestClient, planning_run_id: str) -> None:
    for report_type, (sheet_name, expected_headers, expected_rows) in EXPORT_EXPECTATIONS.items():
        export_response = client.post(
            f"/api/v1/planning-runs/{planning_run_id}/exports",
            json={"report_type": report_type, "file_format": "XLSX"},
        )

        assert export_response.status_code == 201
        export_payload = export_response.json()
        assert export_payload["report_type"] == report_type
        assert export_payload["file_format"] == "XLSX"
        assert export_payload["generated_by_user_display_name"] == "Development Planner"
        assert export_payload["metadata"] == {
            "sheet_names": [sheet_name],
            "sheet_row_counts": {sheet_name: expected_rows},
        }

        download_response = client.get(export_payload["download_url"])
        assert download_response.status_code == 200
        assert download_response.headers["x-report-export-id"] == export_payload["id"]

        workbook = load_workbook(BytesIO(download_response.content), data_only=True)
        assert workbook.sheetnames == ["Export_Info", sheet_name]
        export_info = {
            str(row[0]): row[1]
            for row in workbook["Export_Info"].iter_rows(min_row=2, values_only=True)
        }
        assert export_info["Report_Type"] == report_type
        assert export_info["PlanningRun_ID"] == planning_run_id
        assert export_info["Upload_File"] == "machine_shop_sample_input.xlsx"
        assert export_info["Planning_Start_Date"] == "2026-04-21"
        assert export_info["Planning_Horizon_Days"] == 7

        worksheet = workbook[sheet_name]
        assert [cell.value for cell in worksheet[1]] == expected_headers
        rows = _worksheet_records(worksheet)
        assert len(rows) == expected_rows
        _assert_export_sentinel_rows(report_type=report_type, rows=rows)

    session_factory = create_session_factory()
    with session_factory() as session:
        assert _model_count(session=session, model=ReportExport, planning_run_id=planning_run_id) == len(
            EXPORT_EXPECTATIONS
        )


def _assert_export_sentinel_rows(*, report_type: str, rows: list[dict[str, object | None]]) -> None:
    if report_type == "MACHINE_LOAD":
        assert _row_by(rows, "Machine_Type", "HBM")["Status"] == "OVERLOADED"
        assert _row_by(rows, "Machine_Type", "VTL")["Status"] == "UNDERUTILIZED"
    elif report_type == "SUBCONTRACT_PLAN":
        assert {row["Recommendation_Type"] for row in rows} == {"BATCH_SUBCONTRACT_OPPORTUNITY"}
        assert {row["Suggested_Vendor_ID"] for row in rows} == {"VEN-1"}
        assert {row["Batch_Candidate_Count"] for row in rows} == {3}
    elif report_type == "VALVE_READINESS":
        assert _row_by(rows, "Valve_ID", "V-100")["Status"] == "AT_RISK"
        assert _row_by(rows, "Valve_ID", "V-300")["Risk_Reason"] == "Data issue"
    elif report_type == "FLOW_BLOCKER":
        assert {row["Blocker_Type"] for row in rows} == {
            "MACHINE_OVERLOAD",
            "BATCH_RISK",
            "MISSING_COMPONENT",
            "VENDOR_OVERLOADED",
        }
        assert _row_by(rows, "Blocker_Type", "MISSING_COMPONENT")["Valve_ID"] == "V-300"
    elif report_type == "DAILY_EXECUTION":
        assert _row_by(rows, "Operation_Name", "VTL finish")["Planned_Action"] == "OK_INTERNAL"
        assert _row_by(rows, "Operation_Name", "HBM finish")["Planned_Action"] == "BATCH_SUBCONTRACT_OPPORTUNITY"
    else:
        raise AssertionError(f"Unexpected report type {report_type}.")


def _staging_counts(*, session, upload_batch_id: str) -> dict[str, int]:  # type: ignore[no-untyped-def]
    return {
        str(sheet_name): int(row_count)
        for sheet_name, row_count in session.execute(
            select(ImportStagingRow.sheet_name, func.count())
            .where(ImportStagingRow.upload_batch_id == upload_batch_id)
            .group_by(ImportStagingRow.sheet_name)
        )
    }


def _model_count(*, session, model, planning_run_id: str) -> int:  # type: ignore[no-untyped-def]
    return int(
        session.scalar(
            select(func.count()).select_from(model).where(model.planning_run_id == planning_run_id)
        )
        or 0
    )


def _worksheet_records(worksheet) -> list[dict[str, object | None]]:  # type: ignore[no-untyped-def]
    headers = [cell.value for cell in worksheet[1]]
    return [
        dict(zip(headers, row_values, strict=True))
        for row_values in worksheet.iter_rows(min_row=2, values_only=True)
    ]


def _row_by(
    rows: list[dict[str, object | None]],
    column: str,
    value: object,
) -> dict[str, object | None]:
    for row in rows:
        if row[column] == value:
            return row
    raise AssertionError(f"Could not find row where {column} == {value!r}. Rows: {rows!r}")
