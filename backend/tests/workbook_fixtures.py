from io import BytesIO
from typing import Any

from openpyxl import Workbook

REQUIRED_SHEETS = ["Valve_Plan", "Component_Status", "Routing_Master", "Machine_Master", "Vendor_Master"]


def workbook_bytes(sheets: dict[str, list[list[Any]]] | None = None, include_extra_sheet: bool = False) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)

    for sheet_name, rows in (sheets or minimal_workbook_rows()).items():
        sheet = workbook.create_sheet(sheet_name)
        for row in rows:
            sheet.append(row)

    if include_extra_sheet:
        extra = workbook.create_sheet("Planner_Notes")
        extra.append(["Ignored_Field"])
        extra.append(["This sheet is not part of V1 import."])

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def minimal_workbook_rows() -> dict[str, list[list[Any]]]:
    return {
        "Valve_Plan": [
            [" Valve   ID ", "Order ID", "Customer", "Dispatch Date", "Assembly_Date", "Value (Cr)"],
            ["V-100", "O-100", "Acme", "2026-05-01", "2026-04-28", 1.25],
        ],
        "Component_Status": [
            ["Valve_ID", "Component", "Qty", "Fabrication Required", "Fabrication_Complete", "Expected Ready Date", "Critical"],
            ["V-100", "Body", 1, "Y", "N", "2026-04-24", "Y"],
        ],
        "Routing_Master": [
            ["Component", "Operation No", "Operation Name", "Machine Type", "Std Total Hrs", "Subcontract Allowed"],
            ["Body", 10, "HBM roughing", "HBM", 8, "Y"],
        ],
        "Machine_Master": [
            ["Machine ID", "Machine Type", "Hours per Day", "Efficiency Percent", "Buffer Days", "Active"],
            ["HBM-1", "HBM", 16, 80, 4, "Y"],
        ],
        "Vendor_Master": [
            ["Vendor ID", "Vendor Name", "Primary Process", "Turnaround Days", "Transport Days Total", "Approved"],
            ["VEN-1", "Vendor One", "HBM", 3, 1, "Y"],
        ],
    }
