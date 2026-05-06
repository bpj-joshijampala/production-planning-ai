from copy import deepcopy
from io import BytesIO

from alembic import command
from alembic.config import Config
from fastapi.testclient import TestClient
from openpyxl import load_workbook
import pytest

from app.core.config import get_settings
from app.main import create_app
from tests.workbook_fixtures import minimal_workbook_rows, workbook_bytes


@pytest.fixture(name="client")
def fixture_client(tmp_path, monkeypatch):  # type: ignore[no-untyped-def]
    database_path = tmp_path / "report_exports_api.sqlite3"
    upload_dir = tmp_path / "uploads"
    export_dir = tmp_path / "exports"

    monkeypatch.setenv("DATABASE_URL", f"sqlite:///{database_path.as_posix()}")
    monkeypatch.setenv("UPLOAD_DIR", upload_dir.as_posix())
    monkeypatch.setenv("EXPORT_DIR", export_dir.as_posix())
    get_settings.cache_clear()

    command.upgrade(Config("alembic.ini"), "head")

    with TestClient(create_app()) as test_client:
        yield test_client

    get_settings.cache_clear()


def test_create_and_download_first_build_report_exports(client: TestClient) -> None:
    planning_run_id = _create_calculated_planning_run(client)

    expected_sheets = {
        "MACHINE_LOAD": "Machine_Load",
        "SUBCONTRACT_PLAN": "Subcontract_Plan",
        "VALVE_READINESS": "Valve_Readiness",
        "FLOW_BLOCKER": "Flow_Blockers",
        "DAILY_EXECUTION": "Daily_Execution",
    }

    for report_type, sheet_name in expected_sheets.items():
        create_response = client.post(
            f"/api/v1/planning-runs/{planning_run_id}/exports",
            json={"report_type": report_type, "file_format": "XLSX"},
        )

        assert create_response.status_code == 201
        payload = create_response.json()
        assert payload["planning_run_id"] == planning_run_id
        assert payload["report_type"] == report_type
        assert payload["file_format"] == "XLSX"
        assert payload["generated_by_user_display_name"] == "Development Planner"
        assert payload["download_url"] == f"/api/v1/exports/{payload['id']}/download"
        assert payload["metadata"]["sheet_names"] == [sheet_name]
        assert payload["metadata"]["sheet_row_counts"][sheet_name] >= 0

        detail_response = client.get(f"/api/v1/exports/{payload['id']}")
        assert detail_response.status_code == 200
        assert detail_response.json()["id"] == payload["id"]
        assert detail_response.json()["generated_by_user_display_name"] == "Development Planner"

        download_response = client.get(payload["download_url"])
        assert download_response.status_code == 200
        assert (
            download_response.headers["content-type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert download_response.headers["x-report-export-id"] == payload["id"]

        workbook = load_workbook(filename=BytesIO(download_response.content))
        assert workbook.sheetnames == ["Export_Info", sheet_name]


def test_create_report_export_rejects_unsupported_format(client: TestClient) -> None:
    planning_run_id = _create_calculated_planning_run(client)

    response = client.post(
        f"/api/v1/planning-runs/{planning_run_id}/exports",
        json={"report_type": "MACHINE_LOAD", "file_format": "CSV"},
    )

    assert response.status_code == 422


def test_list_report_exports_returns_history_and_latest_by_report_type(client: TestClient) -> None:
    planning_run_id = _create_calculated_planning_run(client)

    first_machine_load = client.post(
        f"/api/v1/planning-runs/{planning_run_id}/exports",
        json={"report_type": "MACHINE_LOAD", "file_format": "XLSX"},
    ).json()
    latest_machine_load = client.post(
        f"/api/v1/planning-runs/{planning_run_id}/exports",
        json={"report_type": "MACHINE_LOAD", "file_format": "XLSX"},
    ).json()
    flow_blocker = client.post(
        f"/api/v1/planning-runs/{planning_run_id}/exports",
        json={"report_type": "FLOW_BLOCKER", "file_format": "XLSX"},
    ).json()

    history_response = client.get(f"/api/v1/planning-runs/{planning_run_id}/exports?page=1&page_size=100")
    assert history_response.status_code == 200
    history_payload = history_response.json()
    assert history_payload["total"] == 3
    assert {item["id"] for item in history_payload["items"]} == {
        first_machine_load["id"],
        latest_machine_load["id"],
        flow_blocker["id"],
    }

    latest_response = client.get(
        f"/api/v1/planning-runs/{planning_run_id}/exports?latest_only=true&page=1&page_size=100"
    )
    assert latest_response.status_code == 200
    latest_payload = latest_response.json()
    assert latest_payload["total"] == 2

    latest_by_type = {item["report_type"]: item for item in latest_payload["items"]}
    assert latest_by_type["MACHINE_LOAD"]["id"] == latest_machine_load["id"]
    assert latest_by_type["FLOW_BLOCKER"]["id"] == flow_blocker["id"]
    assert latest_by_type["MACHINE_LOAD"]["download_url"] == (
        f"/api/v1/exports/{latest_machine_load['id']}/download"
    )
    assert latest_by_type["MACHINE_LOAD"]["generated_by_user_display_name"] == "Development Planner"


def test_create_report_export_rejects_future_format_and_report_type(client: TestClient) -> None:
    planning_run_id = _create_calculated_planning_run(client)

    pdf_response = client.post(
        f"/api/v1/planning-runs/{planning_run_id}/exports",
        json={"report_type": "MACHINE_LOAD", "file_format": "PDF"},
    )
    assert pdf_response.status_code == 400
    assert pdf_response.json()["detail"]["code"] == "UNSUPPORTED_EXPORT_FORMAT"

    weekly_response = client.post(
        f"/api/v1/planning-runs/{planning_run_id}/exports",
        json={"report_type": "WEEKLY_PLANNING", "file_format": "XLSX"},
    )
    assert weekly_response.status_code == 400
    assert weekly_response.json()["detail"]["code"] == "UNSUPPORTED_REPORT_TYPE"


def test_create_report_export_returns_structured_error_and_logs_unexpected_failure(
    client: TestClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    planning_run_id = _create_calculated_planning_run(client)
    logged_messages: list[str] = []

    def fail_export(*_args: object, **_kwargs: object) -> None:
        raise RuntimeError("synthetic export failure")

    def capture_exception(message: str, *args: object, **_kwargs: object) -> None:
        logged_messages.append(message % args)

    monkeypatch.setattr("app.api.v1.planning_runs.generate_first_build_report_export", fail_export)
    monkeypatch.setattr("app.api.v1.planning_runs.logger.exception", capture_exception)

    response = client.post(
        f"/api/v1/planning-runs/{planning_run_id}/exports",
        json={"report_type": "MACHINE_LOAD", "file_format": "XLSX"},
    )

    assert response.status_code == 500
    assert response.json()["detail"] == {
        "code": "EXPORT_FAILED",
        "message": "Report export failed. Retry the export or contact support if the generated file is still unavailable.",
    }
    assert logged_messages == [
        f"Report export failed planning_run_id={planning_run_id} report_type=MACHINE_LOAD file_format=XLSX"
    ]


def _create_calculated_planning_run(client: TestClient) -> str:
    upload_response = client.post(
        "/api/v1/uploads",
        files={
            "file": (
                "plan.xlsx",
                workbook_bytes(sheets=_first_build_export_workbook_rows()),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert upload_response.status_code == 201

    planning_run_response = client.post(
        "/api/v1/planning-runs",
        json={
            "upload_batch_id": upload_response.json()["id"],
            "planning_start_date": "2026-04-21",
            "planning_horizon_days": 7,
        },
    )
    assert planning_run_response.status_code == 201

    planning_run_id = str(planning_run_response.json()["id"])
    calculate_response = client.post(f"/api/v1/planning-runs/{planning_run_id}/calculate")
    assert calculate_response.status_code == 200
    return planning_run_id


def _first_build_export_workbook_rows() -> dict[str, list[list[object]]]:
    rows = deepcopy(minimal_workbook_rows())

    rows["Valve_Plan"][0] = [
        "Valve_ID",
        "Order_ID",
        "Customer",
        "Dispatch_Date",
        "Assembly_Date",
        "Value_Cr",
        "Priority",
    ]
    rows["Valve_Plan"][1] = ["V-100", "O-100", "Acme", "2026-05-01", "2026-04-22", 1.25, "A"]
    rows["Valve_Plan"].append(["V-200", "O-200", "Beta", "2026-05-02", "2026-04-24", 0.5, "B"])

    rows["Component_Status"][0] = [
        "Valve_ID",
        "Component_Line_No",
        "Component",
        "Qty",
        "Fabrication_Required",
        "Fabrication_Complete",
        "Expected_Ready_Date",
        "Critical",
        "Ready_Date_Type",
    ]
    rows["Component_Status"][1] = ["V-100", 1, "Body", 1, "N", "Y", "2026-04-21", "Y", "CONFIRMED"]
    rows["Component_Status"].append(["V-200", 1, "Bonnet", 1, "N", "Y", "2026-04-21", "Y", "CONFIRMED"])

    rows["Routing_Master"][0] = [
        "Component",
        "Operation_No",
        "Operation_Name",
        "Machine_Type",
        "Std_Total_Hrs",
        "Subcontract_Allowed",
        "Vendor_Process",
    ]
    rows["Routing_Master"][1] = ["Body", 10, "HBM roughing", "HBM", 8, "Y", "HBM"]
    rows["Routing_Master"].append(["Body", 20, "VTL finish", "VTL", 4, "N", ""])
    rows["Routing_Master"].append(["Bonnet", 10, "HBM finish", "HBM", 8, "Y", "HBM"])

    rows["Machine_Master"][0] = [
        "Machine_ID",
        "Machine_Type",
        "Hours_per_Day",
        "Efficiency_Percent",
        "Buffer_Days",
        "Active",
    ]
    rows["Machine_Master"][1] = ["HBM-1", "HBM", 8, 100, 1, "Y"]
    rows["Machine_Master"].append(["VTL-1", "VTL", 8, 100, 3, "Y"])

    rows["Vendor_Master"][0] = [
        "Vendor_ID",
        "Vendor_Name",
        "Primary_Process",
        "Turnaround_Days",
        "Transport_Days_Total",
        "Capacity_Rating",
        "Reliability",
        "Approved",
    ]
    rows["Vendor_Master"][1] = ["VEN-1", "Vendor One", "HBM", 0, 0, "Medium", "A", "Y"]
    rows["Vendor_Master"].append(["VEN-2", "Vendor Two", "VTL", 2, 1, "Low", "B", "Y"])

    return rows
